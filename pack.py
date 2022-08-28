import openpyxl
wb = openpyxl.load_workbook('pack1.xlsx')
sheet = wb.worksheets[0]
sheet1=wb.create_sheet("s1",1)
for rownum in range(2,sheet.max_row,1):
    customername=sheet.cell(rownum,10).value
    phone=sheet.cell(rownum,11).value
    address=sheet.cell(rownum,13).value
    product=sheet.cell(rownum,5).value
    amount=sheet.cell(rownum,6).value
    sheet1.append([customername,phone,address,product + '*'+ str(amount)])
sheet2=wb.create_sheet("s2")
L=[] #L是收件人資料的列表
for rownum in range(1, sheet1.max_row):
    issame = False
    o1 = [sheet1.cell(rownum,1).value, sheet1.cell(rownum,2).value, sheet1.cell(rownum,3).value]
    # 檢查存不存在L裡 
    for o in L:
        if o == o1: #重複值不加進去
            issame = True
            break
    if not issame:
        L.append(o1)

for rownum in range(1, sheet1.max_row): #比對收件人資料 
    for i in L:
        if sheet1.cell(rownum,1).value == i[0] and sheet1.cell(rownum,2).value == i[1] and sheet1.cell(rownum,3).value == i[2]:
            i.append(sheet1.cell(rownum,4).value) #再加入品編跟數量
            continue

for i in L:
    sheet2.append(i) #寫入工作表sheet2

wb.save("pack4.xlsx")
